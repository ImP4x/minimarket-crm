from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from models.empleado_model import (
    crear_empleado, listar_empleados, obtener_empleado_por_id,
    actualizar_empleado, eliminar_empleado, obtener_empleado_por_documento
)
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

empleado_bp = Blueprint("empleado", __name__)

def _usuario_logueado_y_permiso():
    """Helper: devuelve True si hay sesión y rol válido"""
    if "usuario" not in session:
        return False
    rol = session["usuario"].get("rol")
    return rol in ("administrador", "vendedor")

# 📌 Listar empleados y crear nuevo
@empleado_bp.route("/empleados", methods=["GET", "POST"])
def empleados():
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))

    if request.method == "POST":
        # Crear nuevo empleado
        nro_documento = request.form.get("nro_documento", "").strip()
        nombre = request.form.get("nombre", "").strip()
        apellido = request.form.get("apellido", "").strip()
        edad = request.form.get("edad", "").strip()
        genero = request.form.get("genero", "").strip()
        cargo = request.form.get("cargo", "").strip()
        correo = request.form.get("correo", "").strip()
        nro_contacto = request.form.get("nro_contacto", "").strip()
        estado = request.form.get("estado", "activo").strip()
        observaciones = request.form.get("observaciones", "").strip()

        if not nro_documento or not nombre or not apellido or not edad:
            flash("Documento, nombre, apellido y edad son obligatorios.")
            return redirect(url_for("empleado.empleados"))

        # Verificar si ya existe empleado con ese documento
        existente = obtener_empleado_por_documento(nro_documento)
        if existente:
            flash("Ya existe un empleado con ese número de documento.")
            return redirect(url_for("empleado.empleados"))

        try:
            crear_empleado(nro_documento, nombre, apellido, edad, genero, 
                         cargo, correo, nro_contacto, estado, observaciones)
            flash("Empleado creado correctamente.")
        except Exception as e:
            print("❌ Error crear_empleado:", e)
            flash("Ocurrió un error al crear el empleado.")
        return redirect(url_for("empleado.empleados"))

    # GET: listado con posible búsqueda
    q = request.args.get("q", None)
    lista = listar_empleados(q)
    return render_template("empleados.html", empleados=lista, q=q)

# 📌 Editar empleado
@empleado_bp.route("/empleados/editar/<empleado_id>", methods=["POST"])
def editar_empleado(empleado_id):
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))

    nro_documento = request.form.get("nro_documento", None)
    nombre = request.form.get("nombre", None)
    apellido = request.form.get("apellido", None)
    edad = request.form.get("edad", None)
    genero = request.form.get("genero", None)
    cargo = request.form.get("cargo", None)
    correo = request.form.get("correo", None)
    nro_contacto = request.form.get("nro_contacto", None)
    estado = request.form.get("estado", None)
    observaciones = request.form.get("observaciones", None)

    try:
        resultado = actualizar_empleado(empleado_id, nro_documento, nombre, apellido,
                                       edad, genero, cargo, correo, nro_contacto, 
                                       estado, observaciones)
        if resultado.get("modified_count", 0) > 0:
            flash("Empleado actualizado correctamente.")
        else:
            flash("No se realizaron cambios o el empleado no fue encontrado.")
    except Exception as e:
        print("❌ Error actualizar_empleado:", e)
        flash("Ocurrió un error al actualizar el empleado.")

    return redirect(url_for("empleado.empleados"))

# 📌 Eliminar empleado
@empleado_bp.route("/empleados/eliminar/<empleado_id>", methods=["POST"])
def eliminar_empleado_route(empleado_id):
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))

    try:
        resultado = eliminar_empleado(empleado_id)
        if resultado.get("deleted_count", 0) > 0:
            flash("Empleado eliminado correctamente.")
        else:
            flash("No se pudo eliminar el empleado.")
    except Exception as e:
        print("❌ Error eliminar_empleado:", e)
        flash("Ocurrió un error al eliminar el empleado.")

    return redirect(url_for("empleado.empleados"))

# 📌 Búsqueda avanzada de empleados
@empleado_bp.route("/empleados/buscar")
def buscar_empleados():
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))
    
    q = request.args.get("q", None)
    lista = listar_empleados(q)
    return render_template("buscar_empleados.html", empleados=lista, q=q)

# 📌 Exportar empleados a PDF
@empleado_bp.route("/empleados/exportar/pdf")
def exportar_empleados_pdf():
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))
    
    q = request.args.get("q", None)
    empleados_list = listar_empleados(q)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title = Paragraph("<b>Reporte de Empleados</b>", styles['Title'])
    elements.append(title)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    
    # Tabla
    data = [['ID', 'Documento', 'Nombre', 'Apellido', 'Edad', 'Cargo', 'Estado']]
    for emp in empleados_list:
        data.append([
            str(emp.get('id_empleado', '')),
            emp.get('nro_documento', ''),
            emp.get('nombre', ''),
            emp.get('apellido', ''),
            str(emp.get('edad', '')),
            emp.get('cargo', ''),
            emp.get('estado', '')
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name="empleados.pdf",
        mimetype="application/pdf"
    )

# 📌 Exportar empleados a Excel
@empleado_bp.route("/empleados/exportar/excel")
def exportar_empleados_excel():
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))
    
    q = request.args.get("q", None)
    empleados_list = listar_empleados(q)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"
    
    # Encabezados
    headers = ['ID', 'Documento', 'Nombre', 'Apellido', 'Edad', 'Género', 'Cargo', 
               'Correo', 'Contacto', 'Estado', 'Observaciones', 'Fecha Registro']
    ws.append(headers)
    
    # Estilo encabezados
    header_fill = PatternFill(start_color="165d2a", end_color="165d2a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Datos
    for emp in empleados_list:
        ws.append([
            emp.get('id_empleado', ''),
            emp.get('nro_documento', ''),
            emp.get('nombre', ''),
            emp.get('apellido', ''),
            emp.get('edad', ''),
            emp.get('genero', ''),
            emp.get('cargo', ''),
            emp.get('correo', ''),
            emp.get('nro_contacto', ''),
            emp.get('estado', ''),
            emp.get('observaciones', ''),
            emp.get('fecha_registro', '')[:10] if emp.get('fecha_registro') else ''
        ])
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name="empleados.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
