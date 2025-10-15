from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file, Response
from models.contrato_model import (
    crear_contrato, listar_contratos, obtener_contrato_por_id,
    actualizar_contrato, eliminar_contrato, contar_contratos_empleado
)
from models.empleado_model import listar_empleados, obtener_empleado_por_id
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from PyPDF2 import PdfReader

contrato_bp = Blueprint("contrato", __name__)

def _usuario_logueado_y_permiso():
    """Helper: devuelve True si hay sesión y rol válido"""
    if "usuario" not in session:
        return False
    rol = session["usuario"].get("rol")
    return rol in ("administrador", "vendedor")

# 📌 Listar contratos y crear nuevo
@contrato_bp.route("/contratos", methods=["GET", "POST"])
def contratos():
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))

    if request.method == "POST":
        # Crear nuevo contrato
        empleado_id = request.form.get("empleado_id", "").strip()
        tipo_contrato = request.form.get("tipo_contrato", "").strip()
        fecha_inicio = request.form.get("fecha_inicio", "").strip()
        fecha_fin = request.form.get("fecha_fin", "").strip()
        salario = request.form.get("salario", "").strip()
        cargo = request.form.get("cargo", "").strip()
        observaciones = request.form.get("observaciones", "").strip()

        # Validación: contratos indefinidos NO requieren fecha_fin
        if not empleado_id or not tipo_contrato or not fecha_inicio or not salario or not cargo:
            flash("Empleado, tipo de contrato, fecha inicio, salario y cargo son obligatorios.")
            return redirect(url_for("contrato.contratos"))
        
        # Si es indefinido, fecha_fin debe ser None
        if tipo_contrato == "Indefinido":
            fecha_fin = None

        try:
            contrato_id = crear_contrato(empleado_id, tipo_contrato, fecha_inicio, fecha_fin, 
                                        salario, cargo, observaciones)
            if contrato_id:
                flash("✅ Contrato creado correctamente y PDF generado.")
            else:
                flash("⚠️ Contrato creado pero hubo error al generar el PDF.")
        except Exception as e:
            print("❌ Error crear_contrato:", e)
            flash("Ocurrió un error al crear el contrato.")
        return redirect(url_for("contrato.contratos"))

    # GET: listado con posible búsqueda
    q = request.args.get("q", None)
    lista_contratos = listar_contratos(q)
    lista_empleados = listar_empleados()
    return render_template("contratos.html", contratos=lista_contratos, empleados=lista_empleados, q=q)

# 📌 Editar contrato
@contrato_bp.route("/contratos/editar/<contrato_id>", methods=["POST"])
def editar_contrato(contrato_id):
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))

    tipo_contrato = request.form.get("tipo_contrato", None)
    fecha_inicio = request.form.get("fecha_inicio", None)
    fecha_fin = request.form.get("fecha_fin", None)
    salario = request.form.get("salario", None)
    cargo = request.form.get("cargo", None)
    observaciones = request.form.get("observaciones", None)

    # Si es indefinido, ignorar fecha_fin
    if tipo_contrato == "Indefinido":
        fecha_fin = None

    try:
        resultado = actualizar_contrato(contrato_id, tipo_contrato, fecha_inicio,
                                       fecha_fin, salario, cargo, observaciones)
        if resultado.get("modified_count", 0) > 0:
            flash("Contrato actualizado correctamente.")
        else:
            flash("No se realizaron cambios o el contrato no fue encontrado.")
    except Exception as e:
        print("❌ Error actualizar_contrato:", e)
        flash("Ocurrió un error al actualizar el contrato.")

    return redirect(url_for("contrato.contratos"))

# 📌 Eliminar contrato
@contrato_bp.route("/contratos/eliminar/<contrato_id>", methods=["POST"])
def eliminar_contrato_route(contrato_id):
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))

    try:
        resultado = eliminar_contrato(contrato_id)
        if resultado.get("deleted_count", 0) > 0:
            flash("Contrato eliminado correctamente.")
        else:
            flash("No se pudo eliminar el contrato.")
    except Exception as e:
        print("❌ Error eliminar_contrato:", e)
        flash("Ocurrió un error al eliminar el contrato.")

    return redirect(url_for("contrato.contratos"))

# 📌 Descargar PDF del contrato
@contrato_bp.route("/contratos/descargar-pdf/<contrato_id>")
def descargar_pdf_contrato(contrato_id):
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))
    
    try:
        contrato = obtener_contrato_por_id(contrato_id)
        if not contrato or not contrato.get("pdf_base64"):
            flash("❌ No se encontró el PDF del contrato.")
            return redirect(url_for("contrato.contratos"))
        
        # Decodificar PDF de base64
        pdf_bytes = base64.b64decode(contrato["pdf_base64"])
        buffer = BytesIO(pdf_bytes)
        buffer.seek(0)
        
        filename = f"Contrato_{contrato.get('id_contrato', 'N/A')}_{contrato.get('tipo_contrato', 'contrato')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/pdf"
        )
    except Exception as e:
        print(f"❌ Error descargar_pdf_contrato: {e}")
        flash("Ocurrió un error al descargar el PDF.")
        return redirect(url_for("contrato.contratos"))

# 📌 Convertir PDF a Excel y descargar
@contrato_bp.route("/contratos/descargar-excel/<contrato_id>")
def descargar_excel_contrato(contrato_id):
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))
    
    try:
        contrato = obtener_contrato_por_id(contrato_id)
        if not contrato:
            flash("❌ No se encontró el contrato.")
            return redirect(url_for("contrato.contratos"))
        
        # Obtener datos del empleado
        empleado = obtener_empleado_por_id(contrato.get("empleado_id"))
        
        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contrato"
        
        # Título
        ws.merge_cells('A1:B1')
        titulo_cell = ws['A1']
        titulo_cell.value = f"CONTRATO DE TRABAJO - {contrato.get('tipo_contrato', '').upper()}"
        titulo_cell.font = Font(size=14, bold=True, color="FFFFFF")
        titulo_cell.fill = PatternFill(start_color="165d2a", end_color="165d2a", fill_type="solid")
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Datos del contrato
        datos = [
            ["", ""],  # Espacio
            ["INFORMACIÓN DEL CONTRATO", ""],
            ["Nro. Contrato:", contrato.get('id_contrato', 'N/A')],
            ["Tipo de Contrato:", contrato.get('tipo_contrato', '')],
            ["Cargo:", contrato.get('cargo', '')],
            ["Salario:", f"${contrato.get('salario', 0):,.0f}"],
            ["Fecha Inicio:", contrato.get('fecha_inicio', '')],
            ["Fecha Fin:", contrato.get('fecha_fin', 'Indefinido') if contrato.get('fecha_fin') else 'Indefinido'],
            ["", ""],  # Espacio
            ["INFORMACIÓN DEL EMPLEADO", ""],
            ["Nombre Completo:", f"{empleado.get('nombre', '')} {empleado.get('apellido', '')}" if empleado else "N/A"],
            ["Documento:", empleado.get('nro_documento', 'N/A') if empleado else "N/A"],
            ["Edad:", str(empleado.get('edad', 'N/A')) if empleado else "N/A"],
            ["Correo:", empleado.get('correo', 'N/A') if empleado else "N/A"],
            ["Contacto:", empleado.get('nro_contacto', 'N/A') if empleado else "N/A"],
            ["", ""],  # Espacio
            ["OBSERVACIONES", ""],
            [contrato.get('observaciones', 'Sin observaciones'), ""],
        ]
        
        row_num = 2
        for dato in datos:
            ws[f'A{row_num}'] = dato[0]
            ws[f'B{row_num}'] = dato[1]
            
            # Formato de encabezados de sección
            if dato[0] in ["INFORMACIÓN DEL CONTRATO", "INFORMACIÓN DEL EMPLEADO", "OBSERVACIONES"]:
                ws.merge_cells(f'A{row_num}:B{row_num}')
                cell = ws[f'A{row_num}']
                cell.font = Font(size=12, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2d8a4d", end_color="2d8a4d", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Formato de etiquetas
            elif dato[0] and dato[0] != "":
                ws[f'A{row_num}'].font = Font(bold=True)
                ws[f'A{row_num}'].alignment = Alignment(horizontal="left", vertical="center")
                ws[f'B{row_num}'].alignment = Alignment(horizontal="left", vertical="center")
            
            row_num += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Contrato_{contrato.get('id_contrato', 'N/A')}_Excel.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"❌ Error descargar_excel_contrato: {e}")
        flash("Ocurrió un error al generar el Excel.")
        return redirect(url_for("contrato.contratos"))

# 📌 Ver cuántos contratos tiene un empleado
@contrato_bp.route("/contratos/empleado/<empleado_id>")
def contratos_por_empleado(empleado_id):
    if not _usuario_logueado_y_permiso():
        return redirect(url_for("auth.login"))
    
    try:
        empleado = obtener_empleado_por_id(empleado_id)
        if not empleado:
            flash("❌ Empleado no encontrado.")
            return redirect(url_for("contrato.contratos"))
        
        # Buscar contratos de este empleado
        todos_contratos = listar_contratos()
        contratos_empleado = [c for c in todos_contratos if c.get('empleado_id') == empleado_id]
        cantidad = len(contratos_empleado)
        
        return render_template(
            "contratos_empleado.html", 
            empleado=empleado, 
            contratos=contratos_empleado,
            cantidad=cantidad
        )
    except Exception as e:
        print(f"❌ Error contratos_por_empleado: {e}")
        flash("Ocurrió un error al buscar los contratos del empleado.")
        return redirect(url_for("contrato.contratos"))
